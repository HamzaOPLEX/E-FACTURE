from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from dashboard.models import APP_Created_Facture,APP_Facture_items,APP_Settings
from django.shortcuts import HttpResponse, get_object_or_404, redirect, render
from datetime import datetime
from django.db.models import Sum

def DownloadJVFile(request):
    ########## Data Handling #################
    data = []
    TVA_taux = int(APP_Settings.objects.all().first().Company_TVATAUX)
    this_year = int(datetime.today().year)
    all_factures = APP_Created_Facture.objects.all()
    for facture in  all_factures :
        facture_items = APP_Facture_items
        date = facture.Date
        client = facture.Client_Name
        facture_number = str(facture.number).zfill(3)
        libelle = f'FACTURE N° {facture_number}/{date.strftime("%Y")}'
        HT = facture_items.objects.filter(Date__year=this_year,BelongToFacture=facture).aggregate(Sum('PT'))['PT__sum']
        if not HT:
            HT = 0
        TVA = HT / 100 * TVA_taux
        TTC = HT + TVA
        row = ['VE',date.strftime('%Y-%m-%d'), facture.Client_Name, libelle, TTC, '' ]
        data.append(row)
        row = ['VE', '', '', '', '' ,HT]
        data.append(row)
        row = ['VE', '', '', '', '', TVA]
        data.append(row)
    ########################################

    ########## Global Vars #################
    ref = f'A1:F{len(data)+4}'
    JournalVent = Workbook()
    JournalVent_sheet = JournalVent.active
    JournalVent_sheet.title = 'Journal de vente'
    ########################################

    ########### Styling ##################
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
    ###############################

    # add column headings. NB. these must be strings
    JournalVent_sheet.append(['JOURNAL', 'DATE', 'N COMPTE', 'LIBELLE', 'DEBIT', 'CREDIT'])
    JournalVent_sheet.append(['', '', '', '', 'TTC', ''])
    JournalVent_sheet.append(['', '', '', '', '', 'HT'])
    JournalVent_sheet.append( ['','','','','','TVA'])
    for row in data:
        JournalVent_sheet.append(row)

    ####################### Column Border Handler ############################
    def set_border(JournalVent_sheet, cell_range):
        thick = Side(border_style="thick", color="000000")
        for row in JournalVent_sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thick, left=thick,right=thick, bottom=thick)

    set_border(JournalVent_sheet, ref)
    ##################################################################

    ######### Column Widht Handler #########################
    dim_holder = DimensionHolder(worksheet=JournalVent_sheet)
    for col in range(JournalVent_sheet.min_column, JournalVent_sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(JournalVent_sheet, min=col, max=col, width=35)
    JournalVent_sheet.column_dimensions = dim_holder
    ########################################################

    ############### Table Handler  ########################
    tab = Table(displayName="Table1", ref=ref)
    # Add a default style with striped roJournalVent_sheet and banded columns
    style = TableStyleInfo(name='TableStyleLight11')
    tab.tableStyleInfo = style
    JournalVent_sheet.add_table(tab)
    ########################################################

    ############### Table Text Style  ########################
    rows = range(1,len(data)+5)
    columns = range(1, 7)
    for row in rows:
        for col in columns:
            JournalVent_sheet.cell(row, col).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
    ########################################################

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="JournalVent{this_year}.xls"'

    JournalVent.save(response)
    return response
